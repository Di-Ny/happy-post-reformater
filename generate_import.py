"""
Script de génération du fichier d'import Happy Post à partir des commandes Amazon (PDF).
Filtre uniquement les commandes Belgique.

Règles de poids :
  - Piège Seul (x1) / lot de 1 : 0.31 kg
  - Lot de 2 (x2)              : 0.32 kg
  - Lot de 3 (x3)              : 0.35 kg
  - Plusieurs qtés (mixte)     : 0.34 kg
  - Plusieurs lots de 3 (2x3+) : 0.50 kg

Dimensions fixes : 20 x 15 x 15 cm
Pays départ : BE
"""

import openpyxl
import re
import sys
import os
from datetime import date
from copy import copy

# --- CONFIG ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "templates", "template_import_colis.xlsx")
DIMENSIONS = {"longueur": 20, "largeur": 15, "hauteur": 15}
PAYS_DEPART = "FRANCE MÉTROPOLITAINE"
PAYS_ARRIVEE = "BELGIQUE"

# Expéditeur
EXPEDITEUR = {
    "nom": "DAUY",
    "prenom": "Nicolas",
    "entreprise": "FURGO",
    "adresse": "4 rue basse Soumartre",
    "complement": "",
    "code_postal": "34600",
    "ville": "FAUGERES",
    "province": "",
    "email": "contact@furgo.fr",
    "telephone": "0607245125",
}

# Email destinataire par défaut (Amazon ne l'exporte pas)
DEFAULT_DEST_EMAIL = "contact@furgo.fr"

NATURE_CONTENU = "Vente de marchandise"


def determine_weight(product_description, quantity):
    """Détermine le poids selon le type de lot et la quantité commandée."""
    desc_lower = product_description.lower()

    is_lot3 = "lot de 3" in desc_lower or "(x3" in desc_lower
    is_lot2 = "lot de 2" in desc_lower or "(x2" in desc_lower
    is_single = "seul" in desc_lower or "(x1" in desc_lower

    if quantity > 1:
        if is_lot3:
            return 0.50  # Plusieurs lots de 3
        else:
            return 0.34  # Plusieurs quantités mixte
    elif is_lot3:
        return 0.35
    elif is_lot2:
        return 0.32
    elif is_single:
        return 0.31
    else:
        # Fallback : essayer de deviner via le prix
        return 0.34


def parse_orders_from_pdf_text(pdf_path):
    """
    Parse les commandes depuis le texte extrait du PDF Amazon.
    Retourne une liste de dict avec les infos de chaque commande.
    """
    # On utilise pdfplumber pour extraire le texte
    try:
        import pdfplumber
    except ImportError:
        print("Installation de pdfplumber...")
        os.system(f"{sys.executable} -m pip install pdfplumber -q")
        import pdfplumber

    orders = []

    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() + "\n===PAGE_BREAK===\n"

    # Séparer par commande (chaque commande commence par "Adresse de livraison" ou "Adresse d'expédition")
    # On split par page break et on traite chaque page comme une commande
    pages = full_text.split("===PAGE_BREAK===")

    for page_text in pages:
        page_text = page_text.strip()
        if not page_text or ("Adresse de livraison" not in page_text and "Adresse d'expédition" not in page_text and "Adresse d\u2019expédition" not in page_text):
            continue

        lines = page_text.split("\n")
        order = {}

        # Extraire le bloc adresse (entre "Adresse de livraison/expédition :" et la ligne de tirets)
        addr_started = False
        addr_lines = []
        for line in lines:
            if "Adresse de livraison" in line or "Adresse d'expédition" in line or "Adresse d\u2019expédition" in line:
                addr_started = True
                # Parfois le nom est sur la même ligne après ":"
                after = line.split(":", 1)
                if len(after) > 1 and after[1].strip():
                    addr_lines.append(after[1].strip())
                continue
            if addr_started:
                if line.startswith("---") or line.startswith("===") or "Numéro de la commande" in line:
                    break
                if line.strip():
                    addr_lines.append(line.strip())

        # Parser les lignes d'adresse
        # Format typique :
        # Nom Prénom
        # Rue + numéro
        # [Complément]
        # Ville, Code Postal
        # Belgique / France
        # Phone: xxx
        if not addr_lines:
            continue

        # Chercher le pays
        country = ""
        phone = ""
        name_line = ""
        street = ""
        complement = ""
        city = ""
        zipcode = ""
        province = ""
        order_entreprise = ""

        # Identifier les lignes
        filtered_lines = []
        for line in addr_lines:
            phone_match = re.match(r"Phone\s*:\s*(.+)", line, re.IGNORECASE)
            if phone_match:
                phone = phone_match.group(1).strip()
                continue
            line_lower = line.strip().lower()
            if line_lower in ["belgique", "belgium", "france", "belgië"]:
                country = line.strip()
                continue
            # Parfois "Belgique" est collé à la ville/CP (ex: "AUVELAIS, Belgique 5060")
            if re.search(r'\bbelgi(que|um|ë)\b', line_lower):
                country = "Belgique"
                # Retirer "Belgique" de la ligne et garder le reste
                cleaned = re.sub(r',?\s*Belgi(que|um|ë)\s*', ' ', line, flags=re.IGNORECASE).strip()
                if cleaned:
                    filtered_lines.append(cleaned)
                continue
            if line_lower == "france":
                country = "France"
                continue
            filtered_lines.append(line)

        # Filtrer uniquement Belgique
        if country.lower() not in ["belgique", "belgium", "belgië"]:
            continue

        # Nom = première ligne
        if filtered_lines:
            name_line = filtered_lines[0]

        # Chercher la ligne ville/CP : contient un code postal (4 chiffres pour BE)
        # Formats possibles :
        #   "Ville, CODE_POSTAL"       (Amazon BE)
        #   "CODE_POSTAL Ville"         (Amazon FR)
        #   "Ville, PROVINCE CODE_POSTAL"
        # On évite les lignes de rue qui contiennent un numéro de 4 chiffres
        # (ex: "Chaussée de Wavre 1731") en vérifiant le pattern global
        city_zip_idx = -1
        for i, line in enumerate(filtered_lines[1:], 1):
            stripped = line.strip()
            # Pattern 1 : "Ville, CP" ou "Ville, Province CP" (virgule + CP)
            if re.search(r',\s*.*\b\d{4}\b', stripped):
                city_zip_idx = i
                break
            # Pattern 2 : "CP Ville" (ligne commence par 4 chiffres)
            if re.match(r'^\d{4}\s+\S', stripped):
                city_zip_idx = i
                break
            # Pattern 3 : ligne = juste "Ville CP" sans virgule mais pas une rue
            # (une rue a typiquement un mot comme rue/avenue/chaussée/etc.)
            if re.search(r'\b\d{4}\b', stripped):
                has_street_word = re.search(
                    r'\b(rue|avenue|boulevard|chaussée|chemin|place|allée|impasse|'
                    r'drève|sentier|voie|route|ch\.|av\.|bd\.?|bte)\b',
                    stripped, re.IGNORECASE)
                if not has_street_word:
                    city_zip_idx = i
                    break

        if city_zip_idx > 0:
            # Tout entre le nom et la ville = adresse
            street_lines = filtered_lines[1:city_zip_idx]
            # Concaténer les numéros isolés avec la ligne précédente
            # Ex: "Rue du Pont à Biesmes" + "140" -> "Rue du Pont à Biesmes 140"
            merged_street_lines = []
            for sl in street_lines:
                if re.match(r'^\d+[a-zA-Z]?$', sl.strip()) and merged_street_lines:
                    merged_street_lines[-1] = merged_street_lines[-1] + " " + sl.strip()
                else:
                    merged_street_lines.append(sl)

            if len(merged_street_lines) >= 1:
                street = merged_street_lines[0]
            if len(merged_street_lines) >= 2:
                complement = " ".join(merged_street_lines[1:])

            # Si street est juste un numéro, c'est probablement le n° de rue
            # et il manque le nom de la rue (rare mais possible)
            # Vérifier aussi si "street" ressemble à un nom d'entreprise (pas de numéro)
            # et que le complément est la vraie rue
            if complement and not re.search(r'\d', street):
                # street pourrait être une entreprise, complement la rue
                entreprise_candidate = street
                street = complement
                complement = ""
                order_entreprise = entreprise_candidate

            # Parser ville/CP
            city_zip_line = filtered_lines[city_zip_idx]
            # Extraire le code postal (4 chiffres)
            zip_match = re.search(r'\b(\d{4})\b', city_zip_line)
            if zip_match:
                zipcode = zip_match.group(1)
                # La ville est le reste
                city_raw = city_zip_line.replace(zipcode, "").strip().strip(",").strip()
                city_raw = re.sub(r'\s*,\s*', " ", city_raw).strip()

                # Séparer ville et province si présent
                # Ex: "KEERBERGEN VLAAMS BRABANT" -> ville=KEERBERGEN, province=VLAAMS BRABANT
                # Heuristique : la ville est le premier mot, le reste est la province
                # sauf si c'est un nom composé courant
                province = ""
                known_provinces = ["vlaams brabant", "oost-vlaanderen", "west-vlaanderen",
                                   "limburg", "antwerpen", "hainaut", "namur", "liège",
                                   "luxembourg", "brabant wallon", "brabant flamand"]
                city_lower = city_raw.lower()
                for prov in known_provinces:
                    if prov in city_lower:
                        province = city_raw[city_lower.index(prov):city_lower.index(prov)+len(prov)]
                        city = city_raw[:city_lower.index(prov)].strip().strip(",").strip()
                        # Si la ville est vide après extraction de la province,
                        # c'est que la ville a le même nom que la province (ex: Liège)
                        if not city:
                            city = province
                        break
                else:
                    city = city_raw
        else:
            # Fallback
            if len(filtered_lines) > 1:
                street = filtered_lines[1]
            if len(filtered_lines) > 2:
                city = filtered_lines[2]

        # Parser nom/prénom
        name_parts = name_line.split()
        if len(name_parts) >= 2:
            # Heuristique : si tout est en majuscules, c'est NOM PRENOM
            # Sinon Prénom Nom
            last_name = name_parts[0]
            first_name = " ".join(name_parts[1:])
        elif len(name_parts) == 1:
            last_name = name_parts[0]
            first_name = "M."
        else:
            last_name = name_line
            first_name = "M."

        # Extraire le numéro de commande
        order_num = ""
        for line in lines:
            m = re.search(r"Numéro de la commande\s*:\s*([\d-]+)", line)
            if m:
                order_num = m.group(1)
                break

        # Extraire la quantité et le produit
        # Le produit est souvent sur 2+ lignes après "Quantité Détails..."
        # Ligne 1: "1 VVTrap - Piège à frelons ... 27,50 €"
        # Ligne 2: "(Lot de 2 (x2 TVA comprise"
        quantity = 1
        product = ""
        product_lines = []
        in_product = False

        for i, line in enumerate(lines):
            # Détecter la ligne avec quantité + début produit
            qty_match = re.match(r"^\s*(\d+)\s+(.*(?:VVTrap|Velutina|Pi[eè]ge|frelons).*)", line, re.IGNORECASE)
            if qty_match:
                quantity = int(qty_match.group(1))
                product_lines.append(qty_match.group(2).strip())
                in_product = True
                continue

            if in_product:
                # Concaténer les lignes suivantes jusqu'à SKU/ASIN/État
                if re.match(r"^(SKU|ASIN|État|ID de)", line.strip()):
                    in_product = False
                    continue
                if line.strip():
                    product_lines.append(line.strip())
                    # Arrêter après la ligne identifiant le type de lot
                    if "lot de" in line.lower() or "seul" in line.lower():
                        in_product = False

        product = " ".join(product_lines)

        weight = determine_weight(product, quantity)

        # Corriger les téléphones malformés (+0xxx → 0xxx)
        if phone.startswith("+0"):
            phone = phone[1:]  # Retirer le "+"

        order = {
            "nom": last_name,
            "prenom": first_name,
            "entreprise": order_entreprise,
            "adresse": street,
            "complement": complement,
            "code_postal": zipcode,
            "ville": city,
            "province": province,
            "telephone": phone,
            "email": DEFAULT_DEST_EMAIL,
            "pays": PAYS_ARRIVEE,
            "commande": order_num,
            "produit": product,
            "quantite": quantity,
            "poids": weight,
        }

        orders.append(order)

    return orders


def generate_import_file(orders, output_path):
    """Génère le fichier Excel d'import Happy Post à partir des commandes."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    for idx, order in enumerate(orders):
        row = idx + 3  # Les données commencent à la ligne 3 (ligne 1 = groupes, ligne 2 = headers)

        # Colis
        ws.cell(row=row, column=1, value=PAYS_DEPART)           # pays départ
        ws.cell(row=row, column=2, value=order["pays"])        # pays arrivée
        ws.cell(row=row, column=3, value=order["poids"])       # poids (kg)
        ws.cell(row=row, column=4, value=DIMENSIONS["longueur"])  # longueur
        ws.cell(row=row, column=5, value=DIMENSIONS["largeur"])   # largeur
        ws.cell(row=row, column=6, value=DIMENSIONS["hauteur"])   # hauteur
        ws.cell(row=row, column=7, value=order["commande"])    # référence
        ws.cell(row=row, column=8, value=NATURE_CONTENU)        # nature

        # Adresse expéditeur (colonnes I à R = 9 à 18)
        ws.cell(row=row, column=9, value=EXPEDITEUR["nom"])
        ws.cell(row=row, column=10, value=EXPEDITEUR["prenom"])
        ws.cell(row=row, column=11, value=EXPEDITEUR["entreprise"])
        ws.cell(row=row, column=12, value=EXPEDITEUR["adresse"])
        ws.cell(row=row, column=13, value=EXPEDITEUR["complement"])
        ws.cell(row=row, column=14, value=EXPEDITEUR["code_postal"])
        ws.cell(row=row, column=15, value=EXPEDITEUR["ville"])
        ws.cell(row=row, column=16, value=EXPEDITEUR["province"])
        ws.cell(row=row, column=17, value=EXPEDITEUR["email"])
        ws.cell(row=row, column=18, value=EXPEDITEUR["telephone"])

        # Adresse destinataire (colonnes S à AB = 19 à 28)
        ws.cell(row=row, column=19, value=order["nom"])
        ws.cell(row=row, column=20, value=order["prenom"])
        ws.cell(row=row, column=21, value=order["entreprise"])
        ws.cell(row=row, column=22, value=order["adresse"])
        ws.cell(row=row, column=23, value=order["complement"])
        ws.cell(row=row, column=24, value=order["code_postal"])
        ws.cell(row=row, column=25, value=order["ville"])
        ws.cell(row=row, column=26, value=order["province"])
        ws.cell(row=row, column=27, value=order["email"])
        ws.cell(row=row, column=28, value=order["telephone"])

    wb.save(output_path)
    return len(orders)


def generate_preparation(orders, output_path, folder_date):
    """Génère un bon de préparation au format texte."""
    from collections import Counter

    # Compter par type
    types = Counter()
    label_map = {0.31: "x1", 0.32: "x2", 0.35: "x3", 0.34: "Multi", 0.50: "Multi x3"}
    for o in orders:
        t = label_map.get(o["poids"], "?")
        types[t] += 1

    total_pieges = (
        types.get("x1", 0) * 1
        + types.get("x2", 0) * 2
        + types.get("x3", 0) * 3
        + types.get("Multi", 0) * 4
        + types.get("Multi x3", 0) * 6
    )

    lines = []
    lines.append(f"BON DE PREPARATION - {folder_date}")
    lines.append(f"{'=' * 50}")
    lines.append("")
    lines.append(f"  {len(orders)} colis  |  ~{total_pieges} pièges")
    lines.append("")
    lines.append(f"  {'Type':<20} {'Nb colis':>8}")
    lines.append(f"  {'-' * 30}")
    for label, count in [("Piège Seul (x1)", types.get("x1", 0)),
                         ("Lot de 2 (x2)", types.get("x2", 0)),
                         ("Lot de 3 (x3)", types.get("x3", 0)),
                         ("Multi (x4+)", types.get("Multi", 0)),
                         ("Multi x3 (x6/x9)", types.get("Multi x3", 0))]:
        if count > 0:
            lines.append(f"  {label:<20} {count:>8}")
    lines.append("")
    lines.append(f"{'=' * 50}")
    lines.append(f"{'#':<4} {'Nom':<25} {'Ville':<18} {'CP':<6} {'Type':<8}")
    lines.append(f"{'-' * 65}")
    for i, o in enumerate(orders, 1):
        nom = f"{o['nom']} {o['prenom']}".strip()
        t = label_map.get(o["poids"], "?")
        lines.append(f"{i:<4} {nom:<25} {o['ville']:<18} {o['code_postal']:<6} {t:<8}")
    lines.append(f"{'-' * 65}")
    lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_import.py <fichier_amazon.pdf> [date_dossier]")
        print("  date_dossier : format YYYY-MM-DD (défaut: date du jour)")
        sys.exit(1)

    pdf_path = sys.argv[1]
    if not os.path.exists(pdf_path):
        pdf_path = os.path.join(SCRIPT_DIR, pdf_path)

    folder_date = sys.argv[2] if len(sys.argv) > 2 else date.today().strftime("%Y-%m-%d")
    output_dir = os.path.join(SCRIPT_DIR, folder_date)
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, f"import_happypost_{folder_date}.xlsx")

    print(f"Lecture du PDF : {pdf_path}")
    orders = parse_orders_from_pdf_text(pdf_path)

    if not orders:
        print("Aucune commande Belgique trouvée !")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  {len(orders)} commandes Belgique extraites")
    print(f"{'='*60}\n")

    # Afficher le récapitulatif pour vérification
    print(f"{'#':<3} {'Nom':<25} {'Ville':<20} {'CP':<6} {'Produit':<15} {'Poids'}")
    print("-" * 85)
    for i, o in enumerate(orders, 1):
        nom_complet = f"{o['nom']} {o['prenom']}".strip()
        produit_court = "Lot 3" if o["poids"] == 0.35 else "Lot 2" if o["poids"] == 0.32 else "x1" if o["poids"] == 0.31 else "Multi" if o["poids"] == 0.34 else "Multi x3"
        print(f"{i:<3} {nom_complet:<25} {o['ville']:<20} {o['code_postal']:<6} {produit_court:<15} {o['poids']} kg")

    print(f"\nGénération du fichier : {output_path}")
    count = generate_import_file(orders, output_path)
    print(f"Fichier créé avec {count} colis.")

    # Bon de préparation
    prep_path = os.path.join(output_dir, f"preparation_{folder_date}.txt")
    generate_preparation(orders, prep_path, folder_date)
    print(f"Bon de préparation : {prep_path}")

    print(f"\n>>> VERIFIER le fichier avant import ! <<<")


if __name__ == "__main__":
    main()
